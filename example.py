from openpyxl import load_workbook
from pathlib import Path

book_resource = load_workbook('excel/FIREANT_dscp_Quy.xlsx')

name = 'AAA'
sheet = book_resource[name]
for row in sheet.iter_rows(min_row=1,min_col=1,max_row=sheet.max_row, max_col=sheet.max_column):
    print(row[3].value)
#print(sheet.cell(1,1))
#type = 'Nam'
#if type =='Nam':
#    print('exist')
#else:
#    print('not exist')
#print (book_resource.get_sheet_names)
#if '1AAA' in book_resource.sheetnames:
 #   print('exist')
#else:
 #   print('not exist')
#print(ws.title)
#if 'AAA' in book_resource.get_sheet_names:
 #   print('exist')
#else:
 #   print('not  exist')