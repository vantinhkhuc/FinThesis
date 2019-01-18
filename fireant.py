import requests
from lxml import html
import pandas as pd
import json
from openpyxl import load_workbook
from pathlib import Path

USERNAME = "tungj2ee@gmail.com"
PASSWORD = "2010canhdan"

type = 1
year = 2017
quarter = 4
count = 4
symbol = 'IDI'
LOGIN_URL = "https://www.FIREANT.vn/Account/Login"
# URL = 'https://www.FIREANT_.vn/api/Data/Finance/LastestFinancialReports?symbol=' +symbol+'&type='+str(type)+'&year='+str(year)+'&quarter='+str(quarter)+'&count='+str(count)
URL_ref = 'https://www.FIREANT.vn/App#/dashboard'


def Get_company(s,q,fromyear,toyear):
    session_requests = requests.session()
    # Get login csrf token
    result = session_requests.get(LOGIN_URL)
    tree = html.fromstring(result.text)
    # print(result.text)
    authenticity_token = list(set(tree.xpath("//input[@name='__RequestVerificationToken']/@value")))[0]
    #print(authenticity_token)
    # Create payload
    payload = {
        "Email": USERNAME,
        "Password": PASSWORD,
        "__RequestVerificationToken": authenticity_token
    }
    # head = {'Authorization': 'token {}'.format(authenticity_token)}
    head = {"__RequestVerificationToken": authenticity_token}
    # Perform login
    result = session_requests.post(LOGIN_URL, data=payload, headers=dict(referer=URL_ref))
    tree = html.fromstring(result.text)
    token = list(set(tree.xpath("//body/@ng-init")))[0]
    #print(token)
    l = str(token[5:])
    l = l.split("'")
    #print(l[3])
    header = {'RequestVerificationToken': l[3]}
    #Login xong, get data
    # for s in Symbols:
    URL_nam = 'https://www.FIREANT.vn/api/Data/Finance/YearlyFinancialInfo?symbol=' +s+ '&fromYear='+str(fromyear)+'&toYear='+str(toyear)
    URL_quy = 'https://www.FIREANT.vn/api/Data/Finance/QuarterlyFinancialInfo?symbol='+s+'&fromYear='+str(fromyear)+'&fromQuarter=1&toYear='+str(toyear)+'&toQuarter=4'
    URL = ''
    if q == 'Q':
        URL = URL_quy
        f = 'FIREANT_/' + s + '_quy.csv'
    else:
        URL = URL_nam
        f = 'FIREANT_/' + s + '_nam.csv'
    result = session_requests.get(URL, headers = header)
    d = json.loads(result.text)
    if(not d):
        print('error ' + s)
        return pd.DataFrame()
    data = pd.DataFrame()
    for i in d:
        data = data.append(i,ignore_index=True)
    # print(data)
    # exit()
    col_list = list(data)
    col_list.remove("Symbol")
    col_list.remove("Year")
    col_list.insert(0,'Symbol')
    col_list.insert(1, 'Year')
    if q == 'Q':
        col_list.remove('Quarter')
        col_list.insert(2,'Quarter')
    data = data.reindex(columns=col_list)
    if q == 'Q':
        data = data.sort_values(['Year','Quarter'], ascending = [0,0])
    else:
        data = data.sort_values('Year', ascending=0)
    # print(col_list)
    # exit()
    # data.to_csv(f, index=False, header=True,encoding='utf-8-sig')
    return data


def run():
    nhomcp = 'dscp'
    ds = pd.read_csv(nhomcp + '.csv')

    my_file = Path('excel/FIREANT_' + nhomcp + '_Nam.xlsx')
    if my_file.is_file():
        book_nam = load_workbook('excel/FIREANT_' + nhomcp + '_Nam.xlsx')
        writer_nam = pd.ExcelWriter('excel/FIREANT_' + nhomcp + '_Nam.xlsx', engine='openpyxl')
        writer_nam.book = book_nam
        writer_nam.sheets = dict((ws.title, ws) for ws in book_nam.worksheets)
    else:
        writer_nam = pd.ExcelWriter('excel/FIREANT_' + nhomcp + '_Nam.xlsx', engine='openpyxl')

    my_file = Path('excel/FIREANT_' + nhomcp + '_Nam.xlsx')
    if my_file.is_file():
        book_quy = load_workbook('excel/FIREANT_' + nhomcp + '_Quy.xlsx')
        writer_quy = pd.ExcelWriter('excel/FIREANT_' + nhomcp + '_Quy.xlsx', engine='openpyxl')
        writer_quy.book = book_quy
        writer_quy.sheets = dict((ws.title, ws) for ws in book_quy.worksheets)
    else:
        writer_quy = pd.ExcelWriter('excel/FIREANT_' + nhomcp + '_Quy.xlsx', engine='openpyxl')

    for i in ds:
        print('Lay du lieu ',i,' theo nam')
        data = Get_company(i, 'Y', 2017, 2018)
        data.to_excel(writer_nam, sheet_name=i, encoding='utf-8-sig')
        print('Lay du lieu ',i,' theo quy')
        data = Get_company(i, 'Q', 2017, 2018)
        data.to_excel(writer_quy, sheet_name=i, encoding='utf-8-sig')
    writer_nam.save()
    writer_quy.save()
    print('xong')

run()
