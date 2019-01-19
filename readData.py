
import pandas as pd
from openpyxl import load_workbook

from pathlib import Path



def classifyData(type):

    file_source = Path('excel/FIREANT_dscp_' + type + '.xlsx')
    file_result = Path('excel/clsf_' + type + '.xlsx')

    writer_result = pd.ExcelWriter('excel/clsf_' + type + '.xlsx', engine='openpyxl')
    if file_source.is_file():
        book_resource = load_workbook('excel/FIREANT_dscp_' + type + '.xlsx')

        book_result = writer_result.book
        for ws in book_resource.worksheets:

            for row in ws.rows:
                if type == 'Nam':
                    name_sheet_active = row[3].value
                    print(name_sheet_active)
                else:
                    name_sheet_active = row[3].value+'_'+row[4].value
                    print(name_sheet_active)

                if name_sheet_active not in book_result.sheetnames:                    
                    book_result.create_sheet(name_sheet_active)
                    #book_result.sheetnames(name_sheet_active)

                ws_result = book_result[name_sheet_active]
                ws_result.append(row)
        #writer_result.book = book_result
        writer_result.save()
    else:
        print('du lieu dau!!!!')

    print('xong')

classifyData('Quy')